"""Page-sampled ground-truth accuracy audit across pipeline runs.

Workflow
--------
1. `sample`  - draw a deterministic, cluster-stratified sample of pages per
   corpus document and emit plain-text "annotation packets" (raw PDF page
   text + look-back context).  A human/LLM annotator reads the packets and
   writes truth JSONs (accuracy_audit/truth/<doc_key>.json) BLIND to any
   run output.
2. `score`   - compare each run's fields_codegen export against the truth
   on the sampled pages, compute precision/recall/form accuracy/coverage,
   and write an Excel workbook into each run's output folder plus a
   consolidated scored.json for reporting.

Truth JSON schema (one file per document)::

    {"doc_key": "...",
     "pages": [{"page": 133,                # 1-based
                "form_name": "Adverse Event",
                "form_source": "printed" | "carryover" | "none",
                "fields": ["AE Term", "Start Date", ...],   # may be []
                "notes": "optional"}]}

Annotation rules (also embedded in the Excel "Method" sheet):
- A field is a data-entry prompt: a question, an input label, or a row
  prompt in a grid/matrix. One row prompt = one field.
- Option values (Yes/No/checkbox choices), section headers, instructions,
  page furniture and OID bracket annotations are NOT fields.
- form_name is the form title governing the page: printed on the page,
  carried over from a preceding title page, or "none" when undeterminable.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import random
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from common import doc_key  # noqa: E402

PROTO = os.path.dirname(HERE)
BASE = os.path.dirname(os.path.dirname(PROTO))
CORPUS = os.path.join(BASE, "data", "crf_forms")
CLI_OUT = os.path.join(PROTO, "data", "outputs", "out")
AUDIT = os.path.join(PROTO, "accuracy_audit")

SAMPLE_SEED = 42
PAGES_PER_DOC = 10
LOOKBACK_PAGES = 40
PAGE_TEXT_CAP = 4500


# ---------------------------------------------------------------- sampling

def stratified_sample(clusters: list[dict], n_pages: int, target: int,
                      rng: random.Random) -> list[tuple[int, int]]:
    """Proportional stratified sample -> [(page_1based, cluster_idx)].

    Cluster page lists in clusters.json are 0-based.  Largest-remainder
    allocation keeps the sample proportional while guaranteeing the total.
    """
    sizes = [len(c.get("pages", [])) for c in clusters]
    total = sum(sizes)
    if total == 0:
        return []
    target = min(target, total)
    quotas = [target * s / total for s in sizes]
    alloc = [int(q) for q in quotas]
    remainders = sorted(range(len(sizes)), key=lambda i: quotas[i] - alloc[i],
                        reverse=True)
    short = target - sum(alloc)
    for i in remainders[:short]:
        alloc[i] += 1
    for i, a in enumerate(alloc):          # clamp to cluster size
        alloc[i] = min(a, sizes[i])
    # top up if clamping lost slots
    deficit = target - sum(alloc)
    if deficit > 0:
        for i in sorted(range(len(sizes)), key=lambda i: sizes[i] - alloc[i],
                        reverse=True):
            take = min(deficit, sizes[i] - alloc[i])
            alloc[i] += take
            deficit -= take
            if deficit == 0:
                break
    out: list[tuple[int, int]] = []
    for ci, c in enumerate(clusters):
        pages = sorted(c.get("pages", []))
        take = alloc[ci]
        if take <= 0:
            continue
        picked = rng.sample(pages, take)
        out.extend((p + 1, ci) for p in picked)
    out.sort()
    return out


def first_nonempty_line(page_text: str) -> str:
    for ln in page_text.splitlines():
        s = ln.strip()
        if s:
            return s[:110]
    return ""


def cmd_sample(_args: argparse.Namespace) -> None:
    import fitz

    os.makedirs(os.path.join(AUDIT, "packets"), exist_ok=True)
    os.makedirs(os.path.join(AUDIT, "truth"), exist_ok=True)
    manifest: list[dict] = []
    for fname in sorted(os.listdir(CORPUS)):
        if not fname.lower().endswith(".pdf"):
            continue
        key = doc_key(fname)
        cl_path = os.path.join(CLI_OUT, key, "clusters.json")
        if not os.path.isfile(cl_path):
            print(f"skip {key}: no clusters.json")
            continue
        with open(cl_path, encoding="utf-8") as f:
            meta = json.load(f)
        rng = random.Random(f"{key}:{SAMPLE_SEED}")
        samples = stratified_sample(meta.get("clusters", []),
                                    meta.get("pages", 0), PAGES_PER_DOC, rng)
        doc = fitz.open(os.path.join(CORPUS, fname))
        try:
            n = doc.page_count
            page_texts = {}
            needed = set()
            for p, _ in samples:
                needed.add(p)
                for q in range(max(1, p - LOOKBACK_PAGES), p):
                    needed.add(q)
            for p in sorted(needed):
                page_texts[p] = doc[p - 1].get_text("text")
            toc = doc.get_toc()
        finally:
            doc.close()

        lines: list[str] = []
        lines.append(f"DOCUMENT: {fname}")
        lines.append(f"doc_key: {key}")
        lines.append(f"pages: {n} | text_layer_pct: {meta.get('text_layer_pct')}")
        lines.append(f"sampled pages (1-based): {[p for p, _ in samples]}")
        if toc:
            lines.append("")
            lines.append("PDF BOOKMARKS (level, title, page):")
            for lvl, title, pg in toc:
                lines.append(f"  L{lvl} p{pg}: {title[:100]}")
        for p, ci in samples:
            lines.append("")
            lines.append(f"{'=' * 70}")
            lines.append(f"=== PAGE {p} (cluster {ci}) ===")
            if not toc:
                lines.append("--- look-back: first line of preceding pages ---")
                for q in range(max(1, p - LOOKBACK_PAGES), p):
                    fl = first_nonempty_line(page_texts.get(q, ""))
                    if fl:
                        lines.append(f"  p{q}: {fl}")
            lines.append("--- page text ---")
            txt = page_texts.get(p, "")
            if len(txt) > PAGE_TEXT_CAP:
                txt = txt[:PAGE_TEXT_CAP] + "\n[... truncated ...]"
            lines.append(txt)
        packet = os.path.join(AUDIT, "packets", f"{key}.txt")
        with open(packet, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        manifest.append({
            "pdf": fname,
            "doc_key": key,
            "dataiku_key": f"crf_forms_{key}",
            "n_pages": n,
            "text_layer_pct": meta.get("text_layer_pct"),
            "samples": [{"page": p, "cluster": ci} for p, ci in samples],
            "has_toc": bool(toc),
        })
        print(f"{key}: sampled {[p for p, _ in samples]} -> {packet}")
    with open(os.path.join(AUDIT, "manifest.json"), "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=1)
    print(f"manifest: {len(manifest)} documents")


# ---------------------------------------------------------------- scoring

RUNS = [
    {
        "name": "sonnet-cli",
        "label": "Claude Sonnet 4.5 (local CLI)",
        "root": os.path.join(PROTO, "data", "outputs", "out"),
        "tag": "claude_4_5_sonnet",
        "dir_prefix": "",
        "summary": "cli_induction_summary_claude_4_5_sonnet.json",
    },
    {
        "name": "gpt52-cli",
        "label": "GPT 5.2 (local CLI)",
        "root": os.path.join(PROTO, "data", "outputs", "out"),
        "tag": "gpt_5_2",
        "dir_prefix": "",
        "summary": "cli_induction_summary_gpt_5_2.json",
    },
    {
        "name": "sonnet-dataiku",
        "label": "Claude Sonnet 4.5 (Dataiku)",
        "root": os.path.join(PROTO, "dataiku_notebook_pipeline", "out_dataiku_sonnet_4_5"),
        "tag": "bedrock_aws_bedrock_us_anthropic_claude_sonnet_4_5_20250929_v1_0",
        "dir_prefix": "crf_forms_",
        "summary": "induction_summary_bedrock_aws_bedrock_us_anthropic_claude_sonnet_4_5_20250929_v1_0.json",
    },
    {
        "name": "gpt52-dataiku",
        "label": "GPT 5.2 (Dataiku)",
        "root": os.path.join(PROTO, "dataiku_notebook_pipeline", "out_dataiku_gpt_5_2"),
        "tag": "azureopenai_azure_openai_nocache_gpt_5_2",
        "dir_prefix": "crf_forms_",
        "summary": "induction_summary_azureopenai_azure_openai_nocache_gpt_5_2.json",
    },
]

FIELD_TOKEN_SORT = 80
FIELD_PARTIAL = 90
FORM_TOKEN_SORT = 80
FORM_PARTIAL = 88

METHOD_TEXT = """\
Page-sampled ground-truth accuracy audit

Sample design
- 10 pages per document, proportional stratified over Stage-0 layout clusters,
  deterministic seed (per-doc RNG 'doc_key:42'). 110 pages across 11 documents.
- Ground truth was annotated by reading the raw PDF page text of every sampled
  page (packets under accuracy_audit/packets/), blind to any run output.

What counts as a field (truth)
- A data-entry prompt: a question, an input label, or a row prompt in a
  grid/matrix. One row prompt = one field. Repeated identical labels on a page
  collapse to one.
- NOT fields: option values (Yes/No/checkbox choices), section or item-group
  headers, instructions/help text, OID bracket annotations, code lists,
  reference tables, page furniture.
- acceptable_extras: real on-page strings a reasonable extractor may emit
  (named answer slots, group headers). Matching one is neither a hit nor a
  false positive.

Matching (rapidfuzz, normalized: casefold, punctuation stripped)
- Field: token_sort_ratio >= 80 or partial_ratio >= 90, greedy one-to-one by
  descending score.
- Form: scored only on pages whose form title is printed on the page
  (form_source=printed); modal extracted form per page; token_sort >= 80,
  partial >= 88, or normalized containment. Study-ID prefixes are accepted.

Metrics
- precision = TP / (TP + FP)  (acceptable-extra matches excluded from both)
- recall (page-strict) = TP / truth fields on sampled pages
- recall (doc-lenient) = truth fields found anywhere in the document's export
  under a fuzzy form+field match - softens page-attribution effects (e.g.
  MAC186 prints each field on a layout page AND a dictionary page).
- false-positive pages: sampled pages with zero truth fields where the run
  extracted unmatched records.

Caveats
- 10 pages/doc is a sample: per-doc numbers carry sampling error (roughly
  +/-10-15 pp on rates); corpus-level micro totals are the steadier signal.
- Truth was annotated by one annotator (LLM) from text dumps; grid-heavy pages
  are hardest and were annotated conservatively.
"""


def norm(s: str) -> str:
    out = []
    for ch in (s or "").casefold():
        out.append(ch if ch.isalnum() else " ")
    return " ".join("".join(out).split())


def field_score(a: str, b: str) -> float:
    from rapidfuzz import fuzz

    na, nb = norm(a), norm(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 100.0
    ts = fuzz.token_sort_ratio(na, nb)
    pr = fuzz.partial_ratio(na, nb) if min(len(na), len(nb)) >= 4 else 0.0
    if ts >= FIELD_TOKEN_SORT or pr >= FIELD_PARTIAL:
        return max(ts, pr)
    return 0.0


def extra_match(extra: str, got: str) -> bool:
    """Lenient matching against the acceptable-extras / code-rows pool.

    Besides the normal fuzzy match, share of one distinctive token (>=6 chars,
    not numeric) is enough - covers truncated OID rows like '21 PANSS115_'
    vs code_row 'PANSS115_RSORRES'.
    """
    if field_score(extra, got) > 0:
        return True
    te = {t for t in norm(extra).split() if len(t) >= 6 and not t.isdigit()}
    tg = {t for t in norm(got).split() if len(t) >= 6 and not t.isdigit()}
    return bool(te & tg)


def form_match(truth_form: str, got_form: str) -> bool:
    from rapidfuzz import fuzz

    nt, ng = norm(truth_form), norm(got_form)
    if not nt or not ng:
        return False
    if nt == ng or nt in ng or ng in nt:
        return True
    return (fuzz.token_sort_ratio(nt, ng) >= FORM_TOKEN_SORT
            or fuzz.partial_ratio(nt, ng) >= FORM_PARTIAL)


def load_export(run: dict, doc_key_: str) -> list[dict] | None:
    d = os.path.join(run["root"], run["dir_prefix"] + doc_key_)
    p = os.path.join(d, f"fields_codegen_{run['tag']}.csv")
    if not os.path.isfile(p):
        return None
    rows = []
    with open(p, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            try:
                pg = int(float(r.get("page") or 0))
            except ValueError:
                pg = 0
            rows.append({"form": (r.get("form_name") or "").strip(),
                         "field": (r.get("field_name") or "").strip(),
                         "page": pg})
    return rows


def greedy_match(truth_fields: list[str], extracted: list[str]) -> tuple[dict, dict]:
    """One-to-one greedy match by descending score.

    Returns (truth_idx -> ext_idx, ext_idx -> truth_idx).
    """
    scored = []
    for ti, t in enumerate(truth_fields):
        for ei, e in enumerate(extracted):
            s = field_score(t, e)
            if s > 0:
                scored.append((s, ti, ei))
    scored.sort(reverse=True)
    t2e: dict = {}
    e2t: dict = {}
    for s, ti, ei in scored:
        if ti in t2e or ei in e2t:
            continue
        t2e[ti] = ei
        e2t[ei] = ti
    return t2e, e2t


def score_doc(truth: dict, export: list[dict] | None) -> dict:
    """Score one run on one document's sampled truth pages."""
    pages_out = []
    verdicts = []
    tp = fp = acceptable = missed = 0
    truth_total = 0
    form_pages = form_ok_pages = 0
    field_pages = field_pages_covered = 0
    fp_pages = 0
    doc_found = 0

    # doc-lenient pool: distinct (field) names anywhere in export
    all_fields = sorted({r["field"] for r in export}) if export else []

    for pg in truth["pages"]:
        pno = pg["page"]
        tfields = pg.get("fields", [])
        extras = list(pg.get("acceptable_extras", [])) + list(pg.get("code_rows", []))
        truth_total += len(tfields)

        recs = [r for r in (export or []) if r["page"] == pno]
        seen = set()
        uniq = []
        for r in recs:
            k = norm(r["field"])
            if k and k not in seen:
                seen.add(k)
                uniq.append(r)

        t2e, e2t = greedy_match(tfields, [r["field"] for r in uniq])
        page_tp = len(t2e)
        page_missed = [tfields[i] for i in range(len(tfields)) if i not in t2e]

        leftover = [i for i in range(len(uniq)) if i not in e2t]
        page_accept = []
        page_fp = []
        for i in leftover:
            if any(extra_match(x, uniq[i]["field"]) for x in extras):
                page_accept.append(uniq[i])
            else:
                page_fp.append(uniq[i])

        # doc-lenient rescue for missed fields
        page_doc_found = sum(
            1 for m in page_missed
            if any(field_score(m, f) > 0 for f in all_fields))

        # form scoring: printed pages with truth form and any extracted rows
        form_scored = False
        form_correct = None
        modal_form = ""
        if uniq:
            counts: dict = {}
            for r in uniq:
                counts[r["form"]] = counts.get(r["form"], 0) + 1
            modal_form = max(counts, key=lambda k: counts[k])
        if pg.get("form_source") == "printed" and pg.get("form_name") and uniq:
            form_scored = True
            form_pages += 1
            form_correct = form_match(pg["form_name"], modal_form)
            if form_correct:
                form_ok_pages += 1

        if tfields:
            field_pages += 1
            if uniq:
                field_pages_covered += 1
        elif page_fp:
            fp_pages += 1

        tp += page_tp
        fp += len(page_fp)
        acceptable += len(page_accept)
        missed += len(page_missed)
        doc_found += page_doc_found

        for ti, ei in t2e.items():
            verdicts.append({"page": pno, "form": uniq[ei]["form"],
                             "field": uniq[ei]["field"], "verdict": "correct",
                             "truth": tfields[ti]})
        for r in page_accept:
            verdicts.append({"page": pno, "form": r["form"], "field": r["field"],
                             "verdict": "acceptable", "truth": ""})
        for r in page_fp:
            verdicts.append({"page": pno, "form": r["form"], "field": r["field"],
                             "verdict": "false_positive", "truth": ""})

        pages_out.append({
            "page": pno,
            "truth_form": pg.get("form_name") or "",
            "form_source": pg.get("form_source", ""),
            "modal_form": modal_form,
            "form_scored": form_scored,
            "form_correct": form_correct,
            "truth_fields": len(tfields),
            "extracted": len(uniq),
            "tp": page_tp,
            "fp": len(page_fp),
            "acceptable": len(page_accept),
            "missed": page_missed,
            "fp_list": [r["field"] for r in page_fp],
            "doc_found_of_missed": page_doc_found,
        })

    prec = tp / (tp + fp) if (tp + fp) else None
    rec = tp / truth_total if truth_total else None
    rec_doc = (tp + doc_found) / truth_total if truth_total else None
    f1 = (2 * prec * rec / (prec + rec)
          if prec is not None and rec is not None and (prec + rec) else None)
    return {
        "pages": pages_out,
        "verdicts": verdicts,
        "truth_fields": truth_total,
        "tp": tp,
        "fp": fp,
        "acceptable": acceptable,
        "missed": missed,
        "doc_found": doc_found,
        "precision": prec,
        "recall_page": rec,
        "recall_doc": rec_doc,
        "f1": f1,
        "form_pages": form_pages,
        "form_ok_pages": form_ok_pages,
        "field_pages": field_pages,
        "field_pages_covered": field_pages_covered,
        "fp_pages": fp_pages,
    }


# ------------------------------------------------- Rave production GT eval

def rave_gt_eval(export: list[dict] | None) -> dict | None:
    """Full-document eval of the Rave book against production digitized.csv."""
    gt_path = os.path.join(
        BASE, "data", "outputs",
        "outputs_from_384_201_00002_annotated_unique_crf", "digitized.csv")
    if export is None or not os.path.isfile(gt_path):
        return None
    with open(gt_path, encoding="utf-8-sig") as f:
        gt = [(r["form_name"].strip(), r["field_name"].strip())
              for r in csv.DictReader(f)
              if (r.get("form_name") or "").strip() and (r.get("field_name") or "").strip()]
    ex = sorted({(r["form"], r["field"]) for r in export})

    from rapidfuzz import fuzz
    scored = []
    for gi, (gf, gfield) in enumerate(gt):
        ngf = norm(gf)
        for ei, (ef, efield) in enumerate(ex):
            if fuzz.partial_ratio(ngf, norm(ef)) < 70:
                continue
            s = field_score(gfield, efield)
            if s > 0:
                scored.append((s, gi, ei))
    scored.sort(reverse=True)
    gm: dict = {}
    em: dict = {}
    for s, gi, ei in scored:
        if gi in gm or ei in em:
            continue
        gm[gi] = ei
        em[ei] = gi
    tp = len(gm)
    prec = tp / len(ex) if ex else None
    rec = tp / len(gt) if gt else None
    return {"gt_pairs": len(gt), "extracted_pairs": len(ex), "matched": tp,
            "precision": prec, "recall": rec}


# ---------------------------------------------------------------- excel

def fmt_pct(x) -> str:
    return "-" if x is None else f"{100 * x:.0f}%"


def write_excel(run: dict, per_doc: dict, rave: dict | None, out_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(bold=True, color="FFFFFF")
    bad_fill = PatternFill("solid", fgColor="FCE4E4")
    ok_fill = PatternFill("solid", fgColor="E6F4EA")

    def header(ws, cols, widths):
        ws.append(cols)
        for c in ws[1]:
            c.fill = head_fill
            c.font = head_font
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        ws.freeze_panes = "A2"

    ws = wb.active
    ws.title = "Summary"
    header(ws, ["document", "status", "sampled pages", "truth fields",
                "matched (TP)", "missed", "false positives", "acceptable extras",
                "precision", "recall (page)", "recall (doc)", "F1",
                "form pages scored", "form pages correct",
                "field pages covered", "FP pages"],
           [46, 16, 13, 11, 12, 8, 13, 15, 10, 12, 12, 8, 15, 15, 17, 9])
    tot = {"truth": 0, "tp": 0, "fp": 0, "acc": 0, "miss": 0, "docf": 0,
           "fpages": 0, "fok": 0, "cov": 0, "covtot": 0, "fppages": 0}
    for dk in sorted(per_doc):
        d = per_doc[dk]
        s = d["score"]
        ws.append([dk, d["status"], len(s["pages"]), s["truth_fields"], s["tp"],
                   s["missed"], s["fp"], s["acceptable"], fmt_pct(s["precision"]),
                   fmt_pct(s["recall_page"]), fmt_pct(s["recall_doc"]),
                   fmt_pct(s["f1"]), s["form_pages"], s["form_ok_pages"],
                   f"{s['field_pages_covered']}/{s['field_pages']}", s["fp_pages"]])
        if d["status"] != "ok" and d["status"] != "ok_with_warnings":
            for c in ws[ws.max_row]:
                c.fill = bad_fill
        tot["truth"] += s["truth_fields"]
        tot["tp"] += s["tp"]
        tot["fp"] += s["fp"]
        tot["acc"] += s["acceptable"]
        tot["miss"] += s["missed"]
        tot["docf"] += s["doc_found"]
        tot["fpages"] += s["form_pages"]
        tot["fok"] += s["form_ok_pages"]
        tot["cov"] += s["field_pages_covered"]
        tot["covtot"] += s["field_pages"]
        tot["fppages"] += s["fp_pages"]
    prec = tot["tp"] / (tot["tp"] + tot["fp"]) if (tot["tp"] + tot["fp"]) else None
    rec = tot["tp"] / tot["truth"] if tot["truth"] else None
    recd = (tot["tp"] + tot["docf"]) / tot["truth"] if tot["truth"] else None
    f1 = 2 * prec * rec / (prec + rec) if prec and rec else None
    ws.append(["CORPUS (micro)", "", "", tot["truth"], tot["tp"], tot["miss"],
               tot["fp"], tot["acc"], fmt_pct(prec), fmt_pct(rec), fmt_pct(recd),
               fmt_pct(f1), tot["fpages"], tot["fok"],
               f"{tot['cov']}/{tot['covtot']}", tot["fppages"]])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = ok_fill

    ws2 = wb.create_sheet("Page detail")
    header(ws2, ["document", "page", "truth form", "form source",
                 "extracted form (modal)", "form ok", "truth fields",
                 "extracted", "TP", "FP", "missed fields", "false positives"],
           [42, 7, 34, 12, 34, 9, 12, 10, 6, 6, 60, 60])
    wrap = Alignment(wrap_text=True, vertical="top")
    for dk in sorted(per_doc):
        for p in per_doc[dk]["score"]["pages"]:
            ws2.append([dk, p["page"], p["truth_form"], p["form_source"],
                        p["modal_form"],
                        "" if p["form_correct"] is None else ("yes" if p["form_correct"] else "NO"),
                        p["truth_fields"], p["extracted"], p["tp"], p["fp"],
                        "; ".join(p["missed"]), "; ".join(p["fp_list"])])
            if p["form_correct"] is False or p["fp"] or p["missed"]:
                for c in ws2[ws2.max_row]:
                    c.alignment = wrap

    ws3 = wb.create_sheet("Record verdicts")
    header(ws3, ["document", "page", "extracted form", "extracted field",
                 "verdict", "matched truth field"], [42, 7, 34, 55, 15, 55])
    for dk in sorted(per_doc):
        for v in per_doc[dk]["score"]["verdicts"]:
            ws3.append([dk, v["page"], v["form"], v["field"], v["verdict"], v["truth"]])
            if v["verdict"] == "false_positive":
                for c in ws3[ws3.max_row]:
                    c.fill = bad_fill

    if rave:
        ws4 = wb.create_sheet("Rave GT eval")
        header(ws4, ["metric", "value"], [46, 18])
        ws4.append(["document", "384-201-00002_Annotated_Unique_CRF_04Nov2024"])
        ws4.append(["ground truth", "production digitized.csv (full document)"])
        ws4.append(["ground-truth pairs", rave["gt_pairs"]])
        ws4.append(["extracted distinct pairs", rave["extracted_pairs"]])
        ws4.append(["matched pairs", rave["matched"]])
        ws4.append(["precision", fmt_pct(rave["precision"])])
        ws4.append(["recall", fmt_pct(rave["recall"])])

    ws5 = wb.create_sheet("Method")
    ws5.column_dimensions["A"].width = 110
    for line in METHOD_TEXT.splitlines():
        ws5.append([line])
        if line and not line.startswith(("-", " ")):
            ws5[f"A{ws5.max_row}"].font = bold

    wb.save(out_path)


def cmd_score(_args: argparse.Namespace) -> None:
    truth_dir = os.path.join(AUDIT, "truth")
    truths = {}
    for fn in sorted(os.listdir(truth_dir)):
        if fn.endswith(".json"):
            with open(os.path.join(truth_dir, fn), encoding="utf-8") as f:
                t = json.load(f)
            truths[t["doc_key"]] = t

    scored_all = {}
    for run in RUNS:
        summary = {}
        sp = os.path.join(run["root"], run["summary"])
        if os.path.isfile(sp):
            with open(sp, encoding="utf-8") as f:
                for r in json.load(f):
                    summary[r["doc"].removeprefix(run["dir_prefix"])] = r.get("status", "?")
        per_doc = {}
        rave = None
        for dk, truth in truths.items():
            export = load_export(run, dk)
            sc = score_doc(truth, export)
            per_doc[dk] = {
                "status": summary.get(dk, "missing" if export is None else "?"),
                "score": sc,
            }
            if export is not None and dk.startswith("384-201-00002"):
                rave = rave_gt_eval(export)
        out_xlsx = os.path.join(run["root"], f"accuracy_audit_{run['tag']}.xlsx")
        write_excel(run, per_doc, rave, out_xlsx)
        print(f"{run['name']}: wrote {out_xlsx}")
        scored_all[run["name"]] = {
            "label": run["label"],
            "root": os.path.relpath(run["root"], PROTO).replace("\\", "/"),
            "tag": run["tag"],
            "rave_gt": rave,
            "docs": {dk: {"status": d["status"],
                          **{k: v for k, v in d["score"].items() if k != "verdicts"}}
                     for dk, d in per_doc.items()},
        }
    with open(os.path.join(AUDIT, "scored.json"), "w", encoding="utf-8") as f:
        json.dump(scored_all, f, indent=1)
    print(f"scored.json: {os.path.join(AUDIT, 'scored.json')}")

    # console recap
    print()
    print(f"{'run':14s} {'prec':>6s} {'rec_pg':>7s} {'rec_doc':>8s} {'TP':>5s} "
          f"{'FP':>4s} {'miss':>5s} {'form_ok':>8s}")
    for name, s in scored_all.items():
        tp = sum(d["tp"] for d in s["docs"].values())
        fp = sum(d["fp"] for d in s["docs"].values())
        miss = sum(d["missed"] for d in s["docs"].values())
        tf = sum(d["truth_fields"] for d in s["docs"].values())
        docf = sum(d["doc_found"] for d in s["docs"].values())
        fpg = sum(d["form_pages"] for d in s["docs"].values())
        fok = sum(d["form_ok_pages"] for d in s["docs"].values())
        prec = tp / (tp + fp) if tp + fp else 0
        print(f"{name:14s} {100 * prec:5.0f}% {100 * tp / tf:6.0f}% "
              f"{100 * (tp + docf) / tf:7.0f}% {tp:5d} {fp:4d} {miss:5d} "
              f"{fok:3d}/{fpg}")


# ---------------------------------------------------------------- report

def cmd_report(_args: argparse.Namespace) -> None:
    """Generate the HTML accuracy report for the two Dataiku runs."""
    from accuracy_report import build_report  # local module, same dir

    truth_dir = os.path.join(AUDIT, "truth")
    truths = {}
    for fn in sorted(os.listdir(truth_dir)):
        if fn.endswith(".json"):
            with open(os.path.join(truth_dir, fn), encoding="utf-8") as f:
                t = json.load(f)
            truths[t["doc_key"]] = t

    detail = {}
    for run in RUNS:
        summary = {}
        sp = os.path.join(run["root"], run["summary"])
        if os.path.isfile(sp):
            with open(sp, encoding="utf-8") as f:
                for r in json.load(f):
                    summary[r["doc"].removeprefix(run["dir_prefix"])] = r.get("status", "?")
        per_doc = {}
        rave = None
        for dk, truth in truths.items():
            export = load_export(run, dk)
            per_doc[dk] = {
                "status": summary.get(dk, "missing" if export is None else "?"),
                "score": score_doc(truth, export),
            }
            if export is not None and dk.startswith("384-201-00002"):
                rave = rave_gt_eval(export)
        detail[run["name"]] = {"run": run, "per_doc": per_doc, "rave": rave}

    out = os.path.join(PROTO, "dataiku_notebook_pipeline", "accuracy_report.html")
    build_report(detail, out)
    print(f"report: {out}")


# ---------------------------------------------------------------- main

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("sample", help="draw page samples and write packets")
    sub.add_parser("score", help="score runs against truth, write Excel per run")
    sub.add_parser("report", help="write HTML accuracy report for Dataiku runs")
    args = ap.parse_args()
    if args.cmd == "sample":
        cmd_sample(args)
    elif args.cmd == "score":
        cmd_score(args)
    elif args.cmd == "report":
        cmd_report(args)


if __name__ == "__main__":
    main()
